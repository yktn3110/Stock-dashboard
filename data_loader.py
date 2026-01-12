from pathlib import Path
import re

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
import requests

from config import EXCEL_PATH


def parse_period_to_order(s: str | None) -> int | None:
    """
    決算期文字列からソート用の数値キーを作る。

    例:
      '24/1Q'   -> 20241
      '2024Q3'  -> 20243
      '2025/4Q' -> 20254
    """
    if s is None:
        return None

    text = str(s)
    match = re.search(r"(\d{2,4}).*?([1-4])Q", text)
    if not match:
        return None

    year = int(match.group(1))
    if year < 100:
        year += 2000

    quarter = int(match.group(2))
    return year * 10 + quarter


@st.cache_data(show_spinner=False)
def load_data(excel_path: Path = EXCEL_PATH):
    """Excel の銘柄一覧・業績を読み込む。"""
    xls = pd.ExcelFile(excel_path)
    df_list = pd.read_excel(xls, "銘柄一覧")
    df_q = pd.read_excel(xls, "業績")

    # IRリンク列がハイパーリンクの場合にターゲットURLを補完し、http(s)以外は無効化
    try:
        # ハイパーリンク取得には read_only=False が必要
        wb = load_workbook(excel_path, read_only=False, data_only=True)
        if "業績" in wb.sheetnames:
            ws = wb["業績"]
            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if "IRリンク" in header:
                col_idx = header.index("IRリンク") + 1
                links: list[str | None] = []

                def normalize(link_val):
                    if not link_val:
                        return None
                    s = str(link_val).strip()
                    if s.startswith("#"):  # シート内リンクは除外
                        return None
                    if not s.lower().startswith(("http://", "https://")):
                        return None
                    return s

                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    cell = row[0]
                    link = None
                    if cell.hyperlink:
                        link = cell.hyperlink.target
                    elif cell.value:
                        link = cell.value
                    links.append(normalize(link))
                if len(links) == len(df_q):
                    df_q["IRリンク"] = links
    except Exception:
        pass

    if "証券コード" in df_list.columns:
        df_list["証券コード"] = df_list["証券コード"].astype(str)
    if "証券コード" in df_q.columns:
        df_q["証券コード"] = df_q["証券コード"].astype(str)

    for col in ["決算期末日", "決算発表日"]:
        if col in df_q.columns:
            df_q[col] = pd.to_datetime(df_q[col], errors="coerce")

    # 決算期のソートキーを事前に計算しておく
    if "決算期" in df_q.columns:
        df_q["期順"] = df_q["決算期"].apply(parse_period_to_order)

    if all(col in df_q.columns for col in ["証券コード", "決算期末日"]):
        df_q = df_q.sort_values(["証券コード", "決算期末日"])

    return df_list, df_q


def update_latest_announcement(excel_path: Path = EXCEL_PATH) -> tuple[bool, str | None]:
    """業績シートの最新決算発表日とYoYを銘柄一覧に反映する。"""
    if not excel_path.exists():
        return False, f"Excel ファイルが見つかりません: {excel_path}"

    try:
        df_q = pd.read_excel(excel_path, "業績")
    except Exception:
        return False, "業績シートの読み込みに失敗しました。"

    required_cols = {"証券コード", "決算発表日"}
    if not required_cols.issubset(df_q.columns):
        return False, "業績シートに必要な列（証券コード/決算発表日）がありません。"

    df_q["証券コード"] = df_q["証券コード"].astype(str)
    df_q["決算発表日"] = pd.to_datetime(df_q["決算発表日"], errors="coerce")
    df_q = df_q.dropna(subset=["証券コード", "決算発表日"])
    if df_q.empty:
        return False, "業績シートに有効な決算発表日がありません。"

    latest_by_code = df_q.groupby("証券コード")["決算発表日"].max()
    df_q["期順"] = df_q["決算期"].apply(parse_period_to_order)

    try:
        wb = load_workbook(excel_path, read_only=False, data_only=True)
    except PermissionError:
        return False, f"Excel を開いているため更新をスキップしました: {excel_path}"
    except Exception:
        return False, "Excel ファイルの読み込みに失敗しました。"

    if "銘柄一覧" not in wb.sheetnames:
        return False, "銘柄一覧シートが見つかりません。"

    ws = wb["銘柄一覧"]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    required_list_cols = {"証券コード", "最新決算発表", "売上高前年比", "営業利益前年比"}
    if not required_list_cols.issubset(header):
        return False, "銘柄一覧シートに必要な列がありません。"

    code_idx = header.index("証券コード") + 1
    latest_idx = header.index("最新決算発表") + 1
    sales_yoy_idx = header.index("売上高前年比") + 1
    op_yoy_idx = header.index("営業利益前年比") + 1

    yoy_by_code: dict[str, tuple[float | None, float | None]] = {}
    for code, latest_date in latest_by_code.items():
        df_code = df_q[df_q["証券コード"] == code].copy()
        df_latest = df_code[df_code["決算発表日"] == latest_date]
        if df_latest.empty:
            yoy_by_code[code] = (None, None)
            continue
        latest_row = df_latest.iloc[0]
        latest_order = parse_period_to_order(latest_row.get("決算期"))
        if latest_order is None:
            yoy_by_code[code] = (None, None)
            continue
        year = latest_order // 10
        prev_order = (year - 1) * 10 + 4
        df_prev = df_code[df_code["期順"] == prev_order]
        if df_prev.empty:
            yoy_by_code[code] = (None, None)
            continue
        prev_row = df_prev.iloc[0]

        def calc_yoy(current, prev):
            if pd.isna(prev) or prev == 0 or pd.isna(current):
                return None
            return (current - prev) / prev * 100

        sales_yoy = calc_yoy(
            latest_row.get("売上高（四半期）"),
            prev_row.get("売上高（四半期）"),
        )
        op_yoy = calc_yoy(
            latest_row.get("営業利益（四半期）"),
            prev_row.get("営業利益（四半期）"),
        )
        yoy_by_code[code] = (sales_yoy, op_yoy)

    updated = 0
    for row in ws.iter_rows(min_row=2):
        code_val = row[code_idx - 1].value
        if code_val is None:
            continue
        code_str = str(code_val)
        if code_str in latest_by_code:
            cell = row[latest_idx - 1]
            cell.value = latest_by_code[code_str].to_pydatetime()
            cell.number_format = "yyyy/mm/dd"
            updated += 1
            sales_yoy, op_yoy = yoy_by_code.get(code_str, (None, None))
            if sales_yoy is not None:
                cell_sales = row[sales_yoy_idx - 1]
                cell_sales.value = round(float(sales_yoy), 1)
                cell_sales.number_format = "0.0"
            if op_yoy is not None:
                cell_op = row[op_yoy_idx - 1]
                cell_op.value = round(float(op_yoy), 1)
                cell_op.number_format = "0.0"

    try:
        wb.save(excel_path)
    except PermissionError:
        return False, f"Excel を開いているため更新をスキップしました: {excel_path}"
    except Exception:
        return False, "Excel ファイルの保存に失敗しました。"

    return True, f"最新決算発表を更新しました（{updated} 件）。"


def normalize_ticker(code: str | int | None) -> str | None:
    if code is None:
        return None
    text = str(code).strip()
    if not text:
        return None
    if text.endswith(".T"):
        return text
    return f"{text}.T"


@st.cache_data(show_spinner=False, ttl=60)
def get_json_data(ticker: str):
    """
    指定されたティッカーのJSONデータを取得する関数。
    """
    url = (
        "https://query1.finance.yahoo.com/v7/finance/chart/"
        f"{ticker}?range=7d&interval=1m&indicators=quote&includeTimestamps=true"
    )
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    try:
        response = requests.get(url, headers=headers, timeout=5)
    except requests.RequestException:
        return None

    if response.status_code != 200:
        return None
    try:
        return response.json()
    except ValueError:
        return None


def get_current_price(ticker: str | None) -> float | None:
    if not ticker:
        return None
    data = get_json_data(ticker)
    if not data:
        return None
    result = data.get("chart", {}).get("result")
    if not result:
        return None
    meta = result[0].get("meta", {})
    price = meta.get("regularMarketPrice")
    if price is None:
        return None
    try:
        return float(price)
    except (TypeError, ValueError):
        return None
